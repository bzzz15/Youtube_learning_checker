import os
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from tkinter.font import Font
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
import yt_dlp as youtube_dl
from collections import Counter
import nltk
from nltk.corpus import stopwords
from nltk.tokenize import word_tokenize
import json

# Download necessary NLTK data
nltk.download('punkt', quiet=True)
nltk.download('stopwords', quiet=True)

ASCII_ART = r"""
 __     __  _______   __    __  _______   __    __  _______   ______  
|  \   |  \|       \ |  \  |  \|       \ |  \  |  \|       \ |      \ 
| $$   | $$| $$$$$$$\| $$  | $$| $$$$$$$\| $$  | $$| $$$$$$$\ \$$$$$$
| $$   | $$| $$  | $$| $$  | $$| $$__| $$| $$  | $$| $$__| $$  | $$   
 \$$\ /  $$| $$  | $$| $$  | $$| $$    $$| $$  | $$| $$    $$  | $$   
  \$$\  $$ | $$  | $$| $$  | $$| $$$$$$$\| $$  | $$| $$$$$$$\  | $$   
   \$$ $$  | $$__/ $$| $$__/ $$| $$  | $$| $$__/ $$| $$  | $$ _| $$_  
    \$$$   | $$    $$ \$$    $$| $$  | $$ \$$    $$| $$  | $$|   $$ \ 
     \$     \$$$$$$$   \$$$$$$  \$$   \$$  \$$$$$$  \$$   \$$ \$$$$$$
"""

def extract_keywords(text, num_keywords=5):
    tokens = word_tokenize(text.lower())
    stop_words = set(stopwords.words('english'))
    tokens = [token for token in tokens if token.isalnum() and token not in stop_words]
    word_freq = Counter(tokens)
    keywords = [word for word, _ in word_freq.most_common(num_keywords)]
    return keywords

def get_transcript(video_id):
    try:
        ydl_opts = {
            'writesubtitles': True,
            'writeautomaticsub': True,
            'subtitleslangs': ['en'],
            'skip_download': True,
            'outtmpl': 'transcript'
        }
        with youtube_dl.YoutubeDL(ydl_opts) as ydl:
            ydl.download([f'https://www.youtube.com/watch?v={video_id}'])

        with open('transcript.en.vtt', 'r', encoding='utf-8') as f:
            transcript = f.read()

        lines = transcript.split('\n')
        cleaned_lines = [line for line in lines if not line.strip().isdigit() and not '-->' in line and line.strip()]
        cleaned_transcript = ' '.join(cleaned_lines)

        os.remove('transcript.en.vtt')

        return cleaned_transcript
    except Exception as e:
        return f"An error occurred: {str(e)}"

class YouTubeLearningTracker:
    def __init__(self, root):
        self.root = root
        self.root.title("YouTube Learning Tracker")
        self.root.geometry("800x600")
        self.style = ttk.Style()
        self.style.theme_use('clam')
        self.setup_fonts()
        self.setup_styles()
        self.setup_initial_ui()

    def setup_fonts(self):
        self.normal_font = Font(family="Arial", size=10)
        self.bold_font = Font(family="Arial", size=10, weight="bold")

    def setup_styles(self):
        styles = {
            'TFrame': {'background': '#e6f0ff'},
            'TButton': {'font': ('Arial', 10), 'background': '#4a7abc', 'foreground': 'white'},
            'TLabel': {'font': ('Arial', 11), 'background': '#e6f0ff'},
            'Header.TLabel': {'font': ('Arial', 14, 'bold')},
            'Treeview': {'font': ('Arial', 10)},
            'Treeview.Heading': {'font': ('Arial', 10, 'bold')}
        }
        for style, config in styles.items():
            self.style.configure(style, **config)

    def setup_initial_ui(self):
        self.clear_window()
        frame = ttk.Frame(self.root, padding="20", style='TFrame')
        frame.pack(fill=tk.BOTH, expand=True)

        ascii_label = tk.Label(frame, text=ASCII_ART, font=('Courier', 8), bg='#e6f0ff')
        ascii_label.pack(pady=10)

        ttk.Label(frame, text="YouTube Learning Tracker", style='Header.TLabel').pack(pady=20)
        ttk.Label(frame, text="Choose your setup:").pack(pady=10)
        ttk.Button(frame, text="Use Default Settings", command=self.use_default_settings).pack(pady=5)
        ttk.Button(frame, text="Choose Custom Path", command=self.specify_custom_settings).pack(pady=5)

    def use_default_settings(self):
        self.file_path = os.path.join("learning_resources", "video_learning_tracker.xlsx")
        self.setup_main_ui()

    def specify_custom_settings(self):
        file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if file_path:
            if not file_path.endswith('.xlsx'):
                messagebox.showerror("Error", "Please select a valid Excel file (.xlsx).")
                return
            self.file_path = file_path
            self.setup_main_ui()
        else:
            messagebox.showwarning("Warning", "No file selected. Using default settings.")
            self.use_default_settings()

    def setup_main_ui(self):
        self.clear_window()
        self.file_system = FileSystem(os.path.dirname(self.file_path))
        self.workbook = self.load_or_create_workbook()

        main_frame = ttk.Frame(self.root, padding="20", style='TFrame')
        main_frame.pack(fill=tk.BOTH, expand=True)

        input_frame = ttk.Frame(main_frame, style='TFrame')
        input_frame.pack(fill=tk.X, pady=10)

        ttk.Label(input_frame, text="YouTube URL:").grid(row=0, column=0, sticky=tk.W, pady=5)
        self.url_entry = ttk.Entry(input_frame, width=50)
        self.url_entry.grid(row=0, column=1, pady=5, padx=5)

        ttk.Label(input_frame, text="Priority:").grid(row=1, column=0, sticky=tk.W, pady=5)
        self.priority_var = tk.StringVar()
        priority_combo = ttk.Combobox(input_frame, textvariable=self.priority_var, values=["High", "Medium", "Low"])
        priority_combo.grid(row=1, column=1, sticky=tk.W, pady=5, padx=5)
        priority_combo.set("Medium")

        ttk.Button(input_frame, text="Add Video", command=self.add_video).grid(row=2, column=0, columnspan=2, pady=10)

        self.tree = ttk.Treeview(main_frame, columns=("Title", "Author", "Duration", "Priority", "Done"), show="headings", selectmode="browse")
        for col in ("Title", "Author", "Duration", "Priority", "Done"):
            self.tree.heading(col, text=col, command=lambda c=col: self.sort_treeview(c))
            self.tree.column(col, width=100)
        self.tree.column("Title", width=200)
        self.tree.pack(fill=tk.BOTH, expand=True, pady=10)

        scrollbar = ttk.Scrollbar(main_frame, orient=tk.VERTICAL, command=self.tree.yview)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.tree.configure(yscrollcommand=scrollbar.set)

        ttk.Button(main_frame, text="Toggle Done Status", command=self.toggle_done).pack(pady=10)
        ttk.Button(main_frame, text="Show Transcript and Topics", command=self.show_transcript_and_topics).pack(pady=10)

        # Add the "Back" button here
        ttk.Button(main_frame, text="Back", command=self.go_back).pack(pady=10)


        self.setup_treeview_styles()
        self.load_existing_data()




    def setup_treeview_styles(self):
        self.tree.tag_configure("high_priority", background="#FFCCCB", font=self.bold_font)
        self.tree.tag_configure("medium_priority", background="#FFFFC1", font=self.normal_font)
        self.tree.tag_configure("low_priority", background="#CCFFCC", font=self.normal_font)

    def clear_window(self):
        for widget in self.root.winfo_children():
            widget.destroy()

    def load_or_create_workbook(self):
        if os.path.exists(self.file_path):
            workbook = load_workbook(self.file_path)
            self.create_missing_sheets(workbook)
            return workbook
        return self.create_new_workbook(self.file_path)

    def create_new_workbook(self, file_path):
        workbook = Workbook()
        for sheet_name in ["Long Videos", "Mid Videos", "Short Videos"]:
            sheet = workbook.create_sheet(title=sheet_name)
            sheet.append(["YouTube URL", "Title", "Author", "Duration", "Priority", "Done"])
        workbook.remove(workbook["Sheet"])
        workbook.save(file_path)
        return workbook

    def create_missing_sheets(self, workbook):
        required_sheets = ["Long Videos", "Mid Videos", "Short Videos"]
        existing_sheets = workbook.sheetnames
        for sheet_name in required_sheets:
            if sheet_name not in existing_sheets:
                sheet = workbook.create_sheet(title=sheet_name)
                sheet.append(["YouTube URL", "Title", "Author", "Duration", "Priority", "Done"])
        workbook.save(self.file_path)

    def load_existing_data(self):
        for sheet_name in ["Long Videos", "Mid Videos", "Short Videos"]:
            sheet = self.workbook[sheet_name]
            for row in sheet.iter_rows(min_row=2, values_only=True):
                self.insert_item_to_treeview(row[1:6])
        self.sort_treeview("Priority")

    def insert_item_to_treeview(self, values):
        item = self.tree.insert("", "end", values=values)
        self.apply_item_style(item, values[3])  # values[3] is the Priority

    def apply_item_style(self, item, priority):
        if priority == "High":
            self.tree.item(item, tags=("high_priority",))
        elif priority == "Medium":
            self.tree.item(item, tags=("medium_priority",))
        else:
            self.tree.item(item, tags=("low_priority",))

    def add_video(self):
        url = self.url_entry.get()
        priority = self.priority_var.get()
        if url and priority:
            try:
                title, author, duration = self.file_system.fetch_youtube_details(url)
                topics = self.file_system.get_video_topics(url)
                if isinstance(topics, list):
                    self.file_system.update_video_data(url, topics, "Not Started")
                else:
                    self.file_system.update_video_data(url, [], "Not Started")
                self.file_system.insert_youtube_link(self.workbook, url, priority)
                self.insert_item_to_treeview((title, author, f"{duration:.2f} hours", priority, "Not Started"))
                self.url_entry.delete(0, tk.END)
                self.sort_treeview("Priority")
                messagebox.showinfo("Success", "Video added successfully!")
            except Exception as e:
                messagebox.showerror("Error", f"Failed to add video: {str(e)}")
        else:
            messagebox.showwarning("Warning", "Please enter a URL and select a priority.")

    def sort_treeview(self, column):
        items = [(self.tree.set(k, column), k) for k in self.tree.get_children('')]
        items.sort(reverse=True)
        for index, (val, k) in enumerate(items):
            self.tree.move(k, '', index)

    def toggle_done(self):
        selected_item = self.tree.selection()
        if selected_item:
            selected_item = selected_item[0]
            current_values = list(self.tree.item(selected_item)['values'])
            new_status = "Completed" if current_values[4] == "Not Started" else "Not Started"
            current_values[4] = new_status
            self.tree.item(selected_item, values=current_values)
            self.apply_item_style(selected_item, current_values[3])
            self.update_excel_status(current_values)

            url = self.get_url_from_title(current_values[0])
            if url:
                video_data = self.file_system.get_video_data(url)
                self.file_system.update_video_data(url, video_data.get('topics', []), new_status)
        else:
            messagebox.showwarning("Warning", "Please select a video to toggle its status.")

    def get_url_from_title(self, title):
        for sheet_name in ["Long Videos", "Mid Videos", "Short Videos"]:
            sheet = self.workbook[sheet_name]
            for row in sheet.iter_rows(min_row=2, max_col=6, values_only=True):
                if row[1] == title:
                    return row[0]
        return None

    def show_transcript_and_topics(self):
        selected_item = self.tree.selection()
        if selected_item:
            selected_item = selected_item[0]
            title = self.tree.item(selected_item)['values'][0]
            url = self.get_url_from_title(title)
            if url:
                transcript = self.file_system.extract_transcript(url)
                video_data = self.file_system.get_video_data(url)
                topics = video_data.get('topics', [])
                if not topics:
                    topics = self.file_system.get_video_topics(url)
                    if isinstance(topics, list):
                        self.file_system.update_video_data(url, topics, video_data.get('status', 'Not Started'))

                self.display_transcript_and_topics_window(transcript, topics)
            else:
                messagebox.showwarning("Warning", "Could not find URL for the selected video.")
        else:
            messagebox.showwarning("Warning", "Please select a video to show its transcript and topics.")

    def display_transcript_and_topics_window(self, transcript, topics):
        info_window = tk.Toplevel(self.root)
        info_window.title("Video Transcript and Topics")
        info_window.geometry("600x500")

        notebook = ttk.Notebook(info_window)
        notebook.pack(expand=True, fill=tk.BOTH)

        transcript_frame = ttk.Frame(notebook)
        notebook.add(transcript_frame, text="Transcript")

        transcript_text = tk.Text(transcript_frame, wrap=tk.WORD, font=("Arial", 10))
        transcript_text.pack(expand=True, fill=tk.BOTH)
        transcript_text.insert(tk.END, transcript)
        transcript_text.config(state=tk.DISABLED)

        transcript_scrollbar = ttk.Scrollbar(transcript_frame, orient=tk.VERTICAL, command=transcript_text.yview)
        transcript_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        transcript_text.config(yscrollcommand=transcript_scrollbar.set)

        topics_frame = ttk.Frame(notebook)
        notebook.add(topics_frame, text="Topics")

        if isinstance(topics, list):
            for topic in topics:
                ttk.Label(topics_frame, text=f"• {topic}", font=("Arial", 12)).pack(anchor=tk.W, padx=10, pady=5)
        else:
            ttk.Label(topics_frame, text=topics, font=("Arial", 12)).pack(anchor=tk.W, padx=10, pady=5)

    def update_excel_status(self, values):
        title = values[0]
        new_status = values[4]
        for sheet_name in ["Long Videos", "Mid Videos", "Short Videos"]:
            sheet = self.workbook[sheet_name]
            for row in sheet.iter_rows(min_row=2, max_col=6, values_only=False):
                if row[1].value == title:
                    row[5].value = new_status
                    self.file_system.save_workbook(self.workbook, self.file_path)
                    return

    def go_back(self):
        self.setup_initial_ui()

    def on_closing(self):
        if messagebox.askokcancel("Quit", "Do you want to save changes before quitting?"):
            self.file_system.save_workbook(self.workbook, self.file_path)
            self.file_system.save_json_data()
        self.root.destroy()

class FileSystem:
    def __init__(self, root_dir):
        self.root_dir = root_dir
        self.json_file = os.path.join(root_dir, "video_data.json")
        os.makedirs(root_dir, exist_ok=True)
        self.load_json_data()

    def load_json_data(self):
        if os.path.exists(self.json_file):
            with open(self.json_file, 'r') as f:
                self.json_data = json.load(f)
        else:
            self.json_data = {}

    def save_json_data(self):
        with open(self.json_file, 'w') as f:
            json.dump(self.json_data, f, indent=4)

    def update_video_data(self, url, topics, status):
        if url not in self.json_data:
            self.json_data[url] = {}
        self.json_data[url]['topics'] = topics
        self.json_data[url]['status'] = status
        self.save_json_data()

    def get_video_data(self, url):
        return self.json_data.get(url, {})

    def extract_transcript(self, url):
        video_id = self.extract_video_id(url)
        if video_id:
            return get_transcript(video_id)
        else:
            return "Invalid YouTube URL"

    def extract_video_id(self, url):
        import re
        video_id_match = re.search(r"(?:v=|\/)([0-9A-Za-z_-]{11}).*", url)
        if video_id_match:
            return video_id_match.group(1)
        return None

    def get_video_topics(self, url):
        transcript = self.extract_transcript(url)
        if transcript.startswith("Transcript not available") or transcript.startswith("An error occurred"):
            return transcript
        return extract_keywords(transcript)

    def fetch_youtube_details(self, url):
        ydl_opts = {'quiet': True, 'no_warnings': True, 'skip_download': True, 'format': 'best'}
        with youtube_dl.YoutubeDL(ydl_opts) as ydl:
            info_dict = ydl.extract_info(url, download=False)
            return (
                info_dict.get('title', 'Unknown Title'),
                info_dict.get('uploader', 'Unknown Author'),
                info_dict.get('duration', 0) / 3600  # Convert to hours
            )

    def insert_youtube_link(self, workbook, link, priority):
        title, author, duration = self.fetch_youtube_details(link)
        sheet_name = "Long Videos" if duration > 4 else "Mid Videos" if duration > 2 else "Short Videos"
        sheet = workbook[sheet_name]
        row_values = [link, title, author, f"{duration:.2f} hours", priority, "Not Started"]
        sheet.append(row_values)
        self.apply_row_color(sheet, sheet.max_row, priority)

    def apply_row_color(self, sheet, row, priority):
        color = {"High": "FF0000", "Medium": "FFFF00", "Low": "00FF00"}[priority]
        fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        for cell in sheet[row]:
            cell.fill = fill

    def save_workbook(self, workbook, file_path):
        workbook.save(filename=file_path)

if __name__ == "__main__":
    root = tk.Tk()
    app = YouTubeLearningTracker(root)
    root.protocol("WM_DELETE_WINDOW", app.on_closing)
    root.mainloop()
